"""Servico de montagem de dataset e exportacao de relatorios (PDF/XLSX)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from sqlalchemy import select

from app.db.session_manager import get_session
from app.models.apuracao import Apuracao
from app.models.apuracao_item import ApuracaoItem
from app.models.competencia import Competencia
from app.models.empresa import Empresa
from app.models.enums import StatusObrigacao
from app.models.lancamento_fiscal import LancamentoFiscal
from app.models.obra import Obra
from app.models.obrigacao_vencimento import ObrigacaoVencimento
from app.utils.formatters import format_brl, format_date_br


@dataclass(frozen=True)
class ReportFilters:
    empresa_id: int | None
    obra_id: int | None
    competencia_id: int | None
    trimestre_ano: int | None
    trimestre_numero: int | None
    visao: str  # POR_OBRA | CONSOLIDADA
    tipo_relatorio: str  # COMPLETO | MEMORIA | COMPOSICAO | EVOLUCAO | VENCIMENTOS


class ReportExportService:
    def build_dataset(self, filters: ReportFilters) -> dict:
        with get_session() as session:
            empresas = list(session.execute(select(Empresa).order_by(Empresa.razao_social.asc())).scalars().all())
            obras = list(session.execute(select(Obra).order_by(Obra.nome.asc())).scalars().all())
            competencias = list(session.execute(select(Competencia).order_by(Competencia.ano.asc(), Competencia.mes.asc())).scalars().all())

            comp_ids = self._resolve_competencias(filters, competencias)

            ap_stmt = select(Apuracao).where(Apuracao.apuracao_valida.is_(True))
            if filters.empresa_id:
                ap_stmt = ap_stmt.where(Apuracao.empresa_id == filters.empresa_id)
            if filters.obra_id:
                ap_stmt = ap_stmt.where(Apuracao.obra_id == filters.obra_id)
            if comp_ids:
                ap_stmt = ap_stmt.where(Apuracao.competencia_id.in_(comp_ids))
            if filters.visao == "CONSOLIDADA":
                ap_stmt = ap_stmt.where(Apuracao.consolidada.is_(True))
            else:
                ap_stmt = ap_stmt.where(Apuracao.consolidada.is_(False))

            apuracoes = list(session.execute(ap_stmt.order_by(Apuracao.created_at.asc())).scalars().all())
            apuracao_ids = [a.id for a in apuracoes]

            itens: list[ApuracaoItem] = []
            if apuracao_ids:
                itens = list(
                    session.execute(
                        select(ApuracaoItem)
                        .where(ApuracaoItem.apuracao_id.in_(apuracao_ids))
                        .order_by(ApuracaoItem.apuracao_id.asc(), ApuracaoItem.ordem.asc())
                    ).scalars().all()
                )

            lan_stmt = select(LancamentoFiscal)
            if filters.empresa_id:
                lan_stmt = lan_stmt.where(LancamentoFiscal.empresa_id == filters.empresa_id)
            if filters.obra_id:
                lan_stmt = lan_stmt.where(LancamentoFiscal.obra_id == filters.obra_id)
            if comp_ids:
                lan_stmt = lan_stmt.where(LancamentoFiscal.competencia_id.in_(comp_ids))
            lancamentos = list(session.execute(lan_stmt).scalars().all())

            obg_stmt = select(ObrigacaoVencimento)
            if filters.empresa_id:
                obg_stmt = obg_stmt.where(ObrigacaoVencimento.empresa_id == filters.empresa_id)
            if filters.obra_id:
                obg_stmt = obg_stmt.where(ObrigacaoVencimento.obra_id == filters.obra_id)
            if comp_ids:
                obg_stmt = obg_stmt.where(ObrigacaoVencimento.competencia_id.in_(comp_ids))
            obrigacoes = list(
                session.execute(obg_stmt.order_by(ObrigacaoVencimento.data_vencimento.asc())).scalars().all()
            )

        empresa_by_id = {e.id: e for e in empresas}
        obra_by_id = {o.id: o for o in obras}
        comp_by_id = {c.id: c for c in competencias}
        ap_by_id = {a.id: a for a in apuracoes}

        memoria_rows: list[dict[str, str]] = []
        composicao_acc: dict[str, dict[str, Decimal]] = {}
        total_itens = Decimal("0.00")

        for item in itens:
            ap = ap_by_id[item.apuracao_id]
            comp = comp_by_id.get(ap.competencia_id)
            empresa = empresa_by_id.get(ap.empresa_id)
            obra = obra_by_id.get(ap.obra_id) if ap.obra_id else None

            valor = Decimal(str(item.valor_calculado))
            base = Decimal(str(item.base_calculo))
            aliq_pct = (Decimal(str(item.aliquota)) * Decimal("100")).quantize(Decimal("0.01"))
            total_itens += valor

            memoria_rows.append(
                {
                    "empresa": empresa.razao_social if empresa else "-",
                    "obra": obra.nome if obra else "Consolidado",
                    "competencia": comp.referencia if comp else "-",
                    "tributo": item.tributo.value,
                    "passo": item.descricao_passo,
                    "base": format_brl(base),
                    "aliquota": f"{str(aliq_pct).replace('.', ',')}%",
                    "valor": format_brl(valor),
                    "gerado_em": format_date_br(ap.created_at),
                    "observacoes": ap.memoria_resumo or "",
                }
            )

            bucket = composicao_acc.setdefault(item.tributo.value, {"base": Decimal("0.00"), "valor": Decimal("0.00")})
            bucket["base"] += base
            bucket["valor"] += valor

        composicao_rows = []
        for tributo, vals in sorted(composicao_acc.items()):
            base = vals["base"]
            valor = vals["valor"]
            aliq = ((valor / base) * Decimal("100")).quantize(Decimal("0.01")) if base > 0 else Decimal("0.00")
            composicao_rows.append(
                {
                    "tributo": tributo,
                    "base": format_brl(base),
                    "aliquota": f"{str(aliq).replace('.', ',')}%",
                    "valor": format_brl(valor),
                }
            )

        evolucao_map: dict[int, dict[str, Decimal]] = {}
        for l in lancamentos:
            ref = evolucao_map.setdefault(l.competencia_id, {"receita": Decimal("0.00"), "impostos": Decimal("0.00")})
            ref["receita"] += Decimal(str(l.receita_bruta))
        for a in apuracoes:
            ref = evolucao_map.setdefault(a.competencia_id, {"receita": Decimal("0.00"), "impostos": Decimal("0.00")})
            ref["impostos"] += Decimal(str(a.total_impostos))

        evolucao_rows = []
        for comp_id in sorted(evolucao_map.keys(), key=lambda x: (comp_by_id[x].ano, comp_by_id[x].mes)):
            comp = comp_by_id[comp_id]
            row = evolucao_map[comp_id]
            evolucao_rows.append(
                {
                    "competencia": comp.referencia,
                    "faturamento": format_brl(row["receita"]),
                    "impostos": format_brl(row["impostos"]),
                }
            )

        status_labels = {
            StatusObrigacao.EM_ABERTO: "Em aberto",
            StatusObrigacao.PAGO: "Pago",
            StatusObrigacao.VENCIDO: "Vencido",
            StatusObrigacao.CANCELADO: "Cancelado",
            StatusObrigacao.NAO_APLICAVEL: "Nao aplicavel",
        }
        vencimento_rows = []
        for o in obrigacoes:
            comp = comp_by_id.get(o.competencia_id)
            emp = empresa_by_id.get(o.empresa_id)
            obra = obra_by_id.get(o.obra_id) if o.obra_id else None
            vencimento_rows.append(
                {
                    "empresa": emp.razao_social if emp else "-",
                    "obra": obra.nome if obra else "Consolidado",
                    "competencia": comp.referencia if comp else "-",
                    "tributo": o.tributo.value,
                    "vencimento": format_date_br(o.data_vencimento),
                    "valor": format_brl(Decimal(str(o.valor))),
                    "status": status_labels.get(o.status, o.status.value),
                    "codigo": o.codigo_receita,
                }
            )

        total_apuracoes = sum((Decimal(str(a.total_impostos)) for a in apuracoes), start=Decimal("0.00"))
        consistency_delta = (total_itens - total_apuracoes).quantize(Decimal("0.01"))

        filtros_txt = self._filters_text(filters, empresa_by_id, obra_by_id, comp_by_id)

        return {
            "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "filters_text": filtros_txt,
            "tipo_relatorio": filters.tipo_relatorio,
            "visao": filters.visao,
            "memoria_rows": memoria_rows,
            "composicao_rows": composicao_rows,
            "evolucao_rows": evolucao_rows,
            "vencimento_rows": vencimento_rows,
            "summary": {
                "apuracoes": str(len(apuracoes)),
                "lancamentos": str(len(lancamentos)),
                "obrigacoes": str(len(obrigacoes)),
                "total_apuracoes": format_brl(total_apuracoes),
                "total_itens": format_brl(total_itens),
                "delta_consistencia": format_brl(consistency_delta),
            },
            "consistency_ok": consistency_delta == Decimal("0.00"),
        }

    def export_pdf(self, dataset: dict, output_path: str) -> str:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise RuntimeError("Pacote reportlab nao encontrado. Instale com: pip install reportlab") from exc

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
        elems = []

        elems.append(Paragraph("Contabase Digital - Relatorio Fiscal", styles["Title"]))
        elems.append(Paragraph(dataset["filters_text"], styles["Normal"]))
        elems.append(Paragraph(f"Gerado em: {dataset['generated_at']}", styles["Normal"]))
        elems.append(Spacer(1, 10))

        def add_table(title: str, headers: list[str], rows: list[list[str]]) -> None:
            elems.append(Paragraph(title, styles["Heading3"]))
            if not rows:
                elems.append(Paragraph("Sem dados para o filtro informado.", styles["Normal"]))
                elems.append(Spacer(1, 8))
                return
            t = Table([headers] + rows, repeatRows=1)
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#335b87")),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f2f6fb"), colors.white]),
                    ]
                )
            )
            elems.append(t)
            elems.append(Spacer(1, 10))

        tipo = dataset.get("tipo_relatorio", "COMPLETO")

        if tipo in ("COMPLETO", "MEMORIA"):
            add_table(
                "Memoria de Calculo",
                ["Empresa", "Obra", "Competencia", "Tributo", "Base", "Aliquota", "Valor", "Gerado em"],
                [
                    [r["empresa"], r["obra"], r["competencia"], r["tributo"], r["base"], r["aliquota"], r["valor"], r["gerado_em"]]
                    for r in dataset["memoria_rows"]
                ],
            )

        if tipo in ("COMPLETO", "COMPOSICAO"):
            add_table(
                "Composicao dos Tributos",
                ["Tributo", "Base", "Aliquota", "Valor"],
                [[r["tributo"], r["base"], r["aliquota"], r["valor"]] for r in dataset["composicao_rows"]],
            )

        if tipo in ("COMPLETO", "EVOLUCAO"):
            add_table(
                "Evolucao Mensal",
                ["Competencia", "Faturamento", "Impostos"],
                [[r["competencia"], r["faturamento"], r["impostos"]] for r in dataset["evolucao_rows"]],
            )

        if tipo in ("COMPLETO", "VENCIMENTOS"):
            add_table(
                "Vencimentos",
                ["Empresa", "Obra", "Competencia", "Tributo", "Vencimento", "Valor", "Status", "Codigo"],
                [
                    [
                        r["empresa"],
                        r["obra"],
                        r["competencia"],
                        r["tributo"],
                        r["vencimento"],
                        r["valor"],
                        r["status"],
                        r["codigo"],
                    ]
                    for r in dataset["vencimento_rows"]
                ],
            )

        summary = dataset["summary"]
        elems.append(Paragraph("Resumo de Consistencia", styles["Heading3"]))
        elems.append(
            Paragraph(
                f"Apuracoes: {summary['apuracoes']} | Lancamentos: {summary['lancamentos']} | Obrigacoes: {summary['obrigacoes']}"
                f" | Total apuracoes: {summary['total_apuracoes']} | Total itens: {summary['total_itens']}"
                f" | Delta: {summary['delta_consistencia']}",
                styles["Normal"],
            )
        )

        doc.build(elems)
        return str(path)

    def export_xlsx(self, dataset: dict, output_path: str) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise RuntimeError("Pacote openpyxl nao encontrado. Instale com: pip install openpyxl") from exc

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumo"

        ws_summary.append(["Contabase Digital - Relatorio Fiscal"])
        ws_summary.append([dataset["filters_text"]])
        ws_summary.append([f"Gerado em: {dataset['generated_at']}"])
        ws_summary.append([])
        for key, value in dataset["summary"].items():
            ws_summary.append([key, value])
        ws_summary["A1"].font = Font(bold=True)

        def add_sheet(name: str, headers: list[str], rows: list[list[str]]) -> None:
            ws = wb.create_sheet(name)
            ws.append(headers)
            for c in range(1, len(headers) + 1):
                ws.cell(1, c).font = Font(bold=True)
            for row in rows:
                ws.append(row)

        tipo = dataset.get("tipo_relatorio", "COMPLETO")
        if tipo in ("COMPLETO", "MEMORIA"):
            add_sheet(
                "Memoria",
                ["Empresa", "Obra", "Competencia", "Tributo", "Passo", "Base", "Aliquota", "Valor", "Gerado em", "Obs"],
                [
                    [
                        r["empresa"],
                        r["obra"],
                        r["competencia"],
                        r["tributo"],
                        r["passo"],
                        r["base"],
                        r["aliquota"],
                        r["valor"],
                        r["gerado_em"],
                        r["observacoes"],
                    ]
                    for r in dataset["memoria_rows"]
                ],
            )

        if tipo in ("COMPLETO", "COMPOSICAO"):
            add_sheet(
                "Composicao",
                ["Tributo", "Base", "Aliquota", "Valor"],
                [[r["tributo"], r["base"], r["aliquota"], r["valor"]] for r in dataset["composicao_rows"]],
            )

        if tipo in ("COMPLETO", "EVOLUCAO"):
            add_sheet(
                "Evolucao",
                ["Competencia", "Faturamento", "Impostos"],
                [[r["competencia"], r["faturamento"], r["impostos"]] for r in dataset["evolucao_rows"]],
            )

        if tipo in ("COMPLETO", "VENCIMENTOS"):
            add_sheet(
                "Vencimentos",
                ["Empresa", "Obra", "Competencia", "Tributo", "Vencimento", "Valor", "Status", "Codigo"],
                [
                    [
                        r["empresa"],
                        r["obra"],
                        r["competencia"],
                        r["tributo"],
                        r["vencimento"],
                        r["valor"],
                        r["status"],
                        r["codigo"],
                    ]
                    for r in dataset["vencimento_rows"]
                ],
            )

        wb.save(path)
        return str(path)

    @staticmethod
    def _resolve_competencias(filters: ReportFilters, competencias: list[Competencia]) -> list[int]:
        if filters.competencia_id:
            return [filters.competencia_id]
        if filters.trimestre_ano and filters.trimestre_numero:
            start = ((filters.trimestre_numero - 1) * 3) + 1
            return [
                c.id
                for c in competencias
                if c.ano == filters.trimestre_ano and start <= c.mes <= start + 2
            ]
        return [c.id for c in competencias]

    @staticmethod
    def _filters_text(
        filters: ReportFilters,
        empresa_by_id: dict[int, Empresa],
        obra_by_id: dict[int, Obra],
        comp_by_id: dict[int, Competencia],
    ) -> str:
        empresa = empresa_by_id.get(filters.empresa_id).razao_social if filters.empresa_id in empresa_by_id else "Todas"
        obra = obra_by_id.get(filters.obra_id).nome if filters.obra_id in obra_by_id else "Todas"
        if filters.competencia_id and filters.competencia_id in comp_by_id:
            periodo = comp_by_id[filters.competencia_id].referencia
        elif filters.trimestre_ano and filters.trimestre_numero:
            periodo = f"{filters.trimestre_numero}o TRI/{filters.trimestre_ano}"
        else:
            periodo = "Todos"

        return (
            f"Empresa: {empresa} | Obra: {obra} | Periodo: {periodo} "
            f"| Visao: {filters.visao} | Tipo: {filters.tipo_relatorio}"
        )
